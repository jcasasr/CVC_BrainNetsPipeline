# Imports

import csv
from datetime import datetime
import numpy as np
import openpyxl
import os
import pandas as pd
import shutil


# Function to rename all files in a folder by keeping the first 8 characters
def rename_files_to_first_8_chars(folder_path):
    # List all files in the folder
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

    # Loop through each file, rename it, and preserve the extension
    for file_name in files:
        # Split the file into the name and extension
        base_name, file_extension = os.path.splitext(file_name)

        # Truncate the base name to the first 8 characters
        new_base_name = base_name[:8]

        # Combine the new base name with the original file extension
        new_file_name = new_base_name + file_extension

        # Full paths for old and new file names
        old_file_path = os.path.join(folder_path, file_name)
        new_file_path = os.path.join(folder_path, new_file_name)

        # Rename the file
        os.rename(old_file_path, new_file_path)
        #print(f"Renamed: {file_name} -> {new_file_name}")

    print(f'All files in {folder_path} renamed succesfully.')


# Function to create a file with corresponding old ID to new ID using the patient info file
def save_ids_to_csv(excel_filename, csv_filename, start_value):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook.active

    # Open a new CSV file for writing
    with open(csv_filename, mode='w', newline='') as csv_file:
        writer = csv.writer(csv_file)

        # Write the header row
        writer.writerow(['ID', 'ID_old'])

        # Iterate through the rows in the sheet
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=start_value):
            col1_value = row[0]  # Patient ID;

            # Format the row_index as a four-digit string with leading zeros
            formatted_id = f"{row_index:04}"

            # Write the values to the CSV file, if they are not None
            writer.writerow([formatted_id, col1_value])


# Function to read a mapping between two columns of a csv
def read_csv_mapping(csv_filename):
    id_to_name_mapping = {}
    with open(csv_filename, mode='r') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) >= 2:  # Ensure there are at least two columns
                file_id = row[0]
                filename = row[1]
                id_to_name_mapping[filename] = file_id
    return id_to_name_mapping


# Function to remove a given suffix from a file name
def remove_suffix_from_filenames(filenames, suffix):
    cleaned_filenames = [filename.replace(suffix, '') for filename in filenames]
    # Some files in GM have a prefix
    cleaned_filenames = [filename.replace('cs', 's') for filename in cleaned_filenames]
    return cleaned_filenames


# Function to change file names according to mapping, also removes given suffix
def rename_files_in_folder_removing_suffix(folder_path, id_to_name_mapping, suffix):
    files = os.listdir(folder_path)
    cleaned_files = remove_suffix_from_filenames(files, suffix)
    renamed_counter = 0

    for original_filename, cleaned_filename in zip(files, cleaned_files):
        if cleaned_filename in id_to_name_mapping:
            new_filename = id_to_name_mapping[cleaned_filename] + '.csv'
            old_file_path = os.path.join(folder_path, original_filename)
            new_file_path = os.path.join(folder_path, new_filename)
            os.rename(old_file_path, new_file_path)
            #print(f"Renamed {original_filename} to {new_filename}")
            renamed_counter += 1
    
    print(f"Renamed {renamed_counter} files.")


# Function to change file names according to mapping
def rename_files_in_folder(folder_path, id_to_name_mapping):
    files = os.listdir(folder_path)
    renamed_counter = 0

    for filename in files:
        # Remove the .csv extension for comparison
        base_filename, ext = os.path.splitext(filename)
        
        # Look up the base filename in the mapping
        if base_filename in id_to_name_mapping:
            new_filename = id_to_name_mapping[base_filename] + ext
            old_file_path = os.path.join(folder_path, filename)
            new_file_path = os.path.join(folder_path, new_filename)
            os.rename(old_file_path, new_file_path)
            renamed_counter += 1
    
    print(f"Renamed {renamed_counter} files.")


# Function to transform a value to date format
def date_format(date_value):
    # If the date exists, format it as a string (e.g., 'YYYY-MM-DD')
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    else:
        return ''


# Function to calculate time period in years given start and end dates
def calculate_duration(start_date_str, end_date_str):
    # Define the date format (adjust this format based on your actual date format in the CSV)
    date_format = '%Y-%m-%d'  # Example format: 'YYYY-MM-DD'
    
    # Check if the start or end date is missing
    if not start_date_str or not end_date_str:
        return 0  # Return 0 if either date is missing
    
    # Parse the start and end dates
    start_date = datetime.strptime(start_date_str, date_format)
    end_date = datetime.strptime(end_date_str, date_format)
    
    # Calculate the duration (end_date - start_date)
    duration_in_days = (end_date - start_date).days
    
    # Return the number of years rounded to 2 decimals
    return round(duration_in_days / 365.25, 2)


# Function to save patient's information from file 'subject_clinical_data.xlsx'
def save_columns_to_csv_BCN(excel_filename, csv_filename):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook.active

    # Open a new CSV file for writing
    with open(csv_filename, mode='w', newline='') as csv_file:
        writer = csv.writer(csv_file)

        # Write the header row
        writer.writerow(['ID', 'origin', 'gender', 'mstype', 'edss', 'dobirth', 'doscan', 'dostart', 'age', 'DD'])
#        writer.writerow(['ID', 'mstype', 'edss', 'age', 'gender'])

        # Iterate through the rows in the sheet
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=0):

            # Format the row_index as a four-digit string with leading zeros
            formatted_id = f"{row_index:04}"

            gender = row[7] if len(row) > 7 else None
            mstype = row[9] if len(row) > 9 else None
            edss = row[11] if len(row) > 12 else None
            dobirth = date_format(row[3]) if len(row) > 3 else None
            doscan = date_format(row[5]) if len(row) > 5 else None
            dostart = date_format(row[4]) if len(row) > 4 else None

            age = calculate_duration(dobirth, doscan)
            DD = calculate_duration(dostart, doscan)

            # Write the values to the CSV file, if they are not None
            writer.writerow([formatted_id, 'BCN', gender, mstype, edss, dobirth, doscan, dostart, age, DD])


# Function to save patient's information from file 'naples2barcelona_multilayer.xlsx'
def save_columns_to_csv_NAP(excel_filename, csv_filename, start_value):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook.active

    # Create a dictionary that maps specific cell content to values
    mapping = {
        'RR': 0,
        'SP': 1,
        'PP': 2,
        None: -1
    }

    # Open a new CSV file for writing
    with open(csv_filename, mode='w', newline='') as csv_file:
        writer = csv.writer(csv_file)

        # Write the header row
        writer.writerow(['ID', 'origin', 'gender', 'mstype', 'edss', 'dobirth', 'doscan', 'dostart', 'age', 'DD'])

        # Iterate through the rows in the sheet
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=start_value):
            
            # ID: Format the row_index as a four-digit string with leading zeros
            formatted_id = f"{row_index:04}"

            # Change mstype to numbers with mapping
            mstype_letters = row[6] if len(row) > 6 else None

            # Get the mapped value based on cell content, default to Unknown if not in the mapping
            mstype = mapping.get(mstype_letters, 'Unknown')

            gender = row[1] if len(row) > 1 else None
            edss = row[9] if len(row) > 9 and row[9] is not None else 0
            dobirth = date_format(row[2]) if len(row) > 2 else None
            doscan = date_format(row[4]) if len(row) > 4 else None
            dostart = date_format(row[7]) if len(row) > 7 else None

            age = calculate_duration(dobirth, doscan)
            DD = calculate_duration(dostart, doscan)

            # Write the values to the CSV file, if they are not None
            writer.writerow([formatted_id, 'NAP', gender, mstype, edss, dobirth, doscan, dostart, age, DD])


# Function to concatenate two csv files
def concatenate_csv(df1, df2, output_file):
    # Concatenate the dataframes row-wise
    concatenated_df = pd.concat([df1, df2])

    # Save the concatenated dataframe to a new CSV file
    concatenated_df.to_csv(output_file, index=False)


# Function to move all files from one folder to another
def move_files_to_folder(source_folder, destination_folder):
    # Ensure destination folder exists, if not create it
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)
    
    moved_count = 0

    # Move files from the first folder to the destination folder
    for file_name in os.listdir(source_folder):
        source_path = os.path.join(source_folder, file_name)
        destination_path = os.path.join(destination_folder, file_name)
        
        if os.path.isfile(source_path):
            shutil.move(source_path, destination_path)
            moved_count += 1
            #print(f"Moved: {file_name} from {source_folder} to {destination_folder}")

    print(f"Files moved successfully: {moved_count}.")
